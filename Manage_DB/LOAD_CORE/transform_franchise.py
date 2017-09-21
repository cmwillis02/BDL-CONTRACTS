import sqlite3
import openpyxl

stage_path= 'C:\Users\cwillis\Desktop\BDL v3.0\Data\Stage\BDL_STAGE.sqlite'
landing_path= 'C:\Users\cwillis\Desktop\BDL v3.0\Data\Landing\ETL_MFL_RAW.sqlite'

stage_conn = sqlite3.connect(stage_path)
stage = stage_conn.cursor()

input_conn = sqlite3.connect(landing_path)
landing = input_conn.cursor()


def reset_db():
    
    stage.execute('''DELETE FROM franchise_fact''')
    #stage.execute(
    #               '''CREATE TABLE IF NOT EXISTS franchise_fact (  ffid INTEGER UNIQUE NOT NULL PRIMARY KEY AUTOINCREMENT,
    #                                                               franchise_id INTEGER NOT NULL,
    #                                                               week_id INTEGER NOT NULL,
    #                                                               opponent_ID INTEGER,
    #                                                               matchup_type TEXT,
    #                                                               result TEXT,
    #                                                               total_score FLOAT)'''
    #                   )
    
    stage_conn.commit()

def matchup_bridge(year):
    
    for week in range(1,17):
        landing.execute(
                            '''SELECT DISTINCT franchise_id, matchup_id, result FROM weeklyresults WHERE year = ? AND week = ?''',(year,week)
                            )   
                        
        results=landing.fetchall()
        for row in results:
            franchise_id = row[0]
            matchup_id = row[1]
            result=row[2]
            
            landing.execute(
                                '''SELECT DISTINCT franchise_id FROM weeklyresults WHERE year=? AND week=? AND matchup_id = ? AND franchise_id <> ?''',(year, week, matchup_id, franchise_id)
                                )
            results=landing.fetchall()
            if len(results)==0:
                opponent_id=None
            else:
                opponent_id=results[0][0]

            
            stage.execute(
                            '''SELECT week_id FROM week_dim WHERE year = ? and week = ?''',(year, week)
                            )
            week_id=stage.fetchone()[0]
            
            if week in range(0,14):
                matchup_type = 'Regular Season'
            elif result == 'BYE':
                matchup_type = 'BYE'
            else:
                matchup_type= 'Ultimate Loser'
            
            print week_id, opponent_id
            
            stage.execute(
                            '''INSERT INTO franchise_fact (franchise_id, week_id, opponent_id, matchup_type, result)
                                                    VALUES(?, ?, ?, ?, ?)''',(franchise_id, week_id, opponent_id, matchup_type, result)
                                )
                                
    stage_conn.commit()
    
def set_playoff_matchups():

    wb = openpyxl.load_workbook('C:\Users\cwillis\Desktop\BDL v3.0\Data\Reference\playoff_reference.xlsx')
    sheet = wb.get_sheet_by_name('playoff matchups')
    for i in range(1,100):
        week_id=sheet.cell(row=i, column=4).value
        franchise_id=sheet.cell(row=i, column=5).value
        
        stage.execute(
                        '''UPDATE franchise_fact SET matchup_type = 'Playoff' WHERE week_id= ? AND franchise_id = ?''',(week_id, franchise_id)
                        )
        
        print week_id, franchise_id

    stage_conn.commit()
    
def add_scores():

    stage.execute(
                        '''SELECT week_id, franchise_id FROM franchise_fact WHERE total_score IS NULL'''
                        )
    for week_id, franchise_id in stage.fetchall():
        
        print week_id, franchise_id
        stage.execute(
                        '''SELECT week_id, franchise_id, sum(score) FROM player_fact WHERE roster_status = 'starter' AND week_id = ? AND franchise_id = ? GROUP BY week_id, franchise_id''',(week_id, franchise_id)
                        )
        
        total_score=round(stage.fetchone()[2],1)
        
        stage.execute(
                        '''UPDATE franchise_fact SET total_score = ? WHERE week_id = ? AND franchise_id = ?''',(total_score,week_id, franchise_id)
                        )
        
        print week_id, franchise_id, total_score
        
    stage_conn.commit()
    

reset_db()   
for year in [2017]:
   matchup_bridge(year)
   
#set_playoff_matchups()
add_scores()
    

 